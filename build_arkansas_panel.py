"""
build_arkansas_panel.py
=======================
Builds the Arkansas DiD panel with FARS coverage through 2023.

HOW TO USE
----------
1.  Download FARS data from NHTSA:
        https://www.nhtsa.gov/file-downloads?p=nhtsa/downloads/FARS/
    For each year 2008-2023:
        click the year folder -> National -> download FARS{YEAR}NationalCSV.zip
        e.g. FARS2021NationalCSV.zip, FARS2022NationalCSV.zip, FARS2023NationalCSV.zip
    Put all zip files in a single folder on your machine.

2.  Set the two paths below:
        FARS_ZIP_DIR  ->  folder containing all the FARS*NationalCSV.zip files
        XLSX_PATH     ->  path to arkansas_county_adjacency_wetdry_panel_2010_2024.xlsx

3.  Run:
        python build_arkansas_panel.py

Requirements
------------
    pip install pandas openpyxl

Outputs (written to OUT_DIR, default = same folder as this script)
-------------------------------------------------------------------
    arkansas_did_panel.csv      main panel, 2008-2023
    transition_summary.csv      one row per wet/dry transition event
    merge_log.txt               full run log
"""

import sys
import zipfile
import pandas as pd
import openpyxl
from pathlib import Path

# ======================================================================
#  SET THESE PATHS BEFORE RUNNING
# ======================================================================
FARS_ZIP_DIR = Path("./fars_zips")          # folder with FARS*NationalCSV.zip files
XLSX_PATH    = Path("arkansas_county_adjacency_wetdry_panel_2010_2024.xlsx")
OUT_DIR      = Path(".")                    # output folder for CSVs and log
# ======================================================================

FARS_YEARS       = list(range(2008, 2024))  # 2008-2023 inclusive
AR_STATE_CODE    = 5                        # Arkansas FIPS state code
ESTIMATION_START = 2008
ESTIMATION_END   = 2023

LOG_LINES = []


# -- Helpers -------------------------------------------------------------------

def log(msg: str = ""):
    print(msg)
    LOG_LINES.append(msg)

def sep(title=""):
    log("=" * 60)
    if title:
        log(title)
        log("=" * 60)

def find_zip(year: int):
    for pattern in [f"*{year}*National*.zip", f"*{year}*.zip"]:
        hits = list(FARS_ZIP_DIR.glob(pattern))
        if hits:
            return hits[0]
    return None


def extract_arkansas(year: int):
    """
    Open the FARS zip for one year, filter to Arkansas,
    and aggregate to county x year.
    Returns a DataFrame or None if the zip is missing.
    """
    zpath = find_zip(year)
    if zpath is None:
        log(f"  [{year}] ZIP NOT FOUND in {FARS_ZIP_DIR} -- skipping")
        return None

    with zipfile.ZipFile(zpath) as zf:
        names = zf.namelist()
        acc_name = next(
            (n for n in names if "accident" in n.lower() and n.lower().endswith(".csv")),
            None,
        )
        if acc_name is None:
            log(f"  [{year}] No accident CSV found in zip. Contents: {names[:10]}")
            return None
        with zf.open(acc_name) as f:
            df = pd.read_csv(f, encoding="latin-1", low_memory=False)

    df.columns = [c.upper().strip() for c in df.columns]

    if "STATE" not in df.columns:
        log(f"  [{year}] ERROR: No STATE column. Found: {list(df.columns)[:20]}")
        return None

    ar = df[df["STATE"] == AR_STATE_CODE].copy()
    log(f"  {year}: {len(ar):,} Arkansas fatal crash records extracted")
    if len(ar) == 0:
        return None

    # Drop unknown county codes (0 = unknown, 999 = not applicable)
    ar["COUNTY"] = pd.to_numeric(ar["COUNTY"], errors="coerce")
    ar = ar[ar["COUNTY"].notna() & (ar["COUNTY"] > 0) & (ar["COUNTY"] < 999)]

    # 5-digit FIPS
    ar["fips"] = "05" + ar["COUNTY"].astype(int).astype(str).str.zfill(3)

    # Alcohol flag: DRUNK_DR = number of drunk drivers involved in crash
    if "DRUNK_DR" in ar.columns:
        ar["alcohol_flag"] = (
            pd.to_numeric(ar["DRUNK_DR"], errors="coerce").fillna(0) > 0
        ).astype(int)
    elif "DR_DRINK" in ar.columns:   # older column name used in some FARS releases
        ar["alcohol_flag"] = (
            pd.to_numeric(ar["DR_DRINK"], errors="coerce").fillna(0) > 0
        ).astype(int)
    else:
        log(f"  [{year}] WARNING: no alcohol column found; alcohol_flag set to 0")
        ar["alcohol_flag"] = 0

    # Fatalities (column name changed across FARS releases)
    fat_col = next((c for c in ["FATALS", "FATALITIES"] if c in ar.columns), None)
    ar["fatals"] = (
        pd.to_numeric(ar[fat_col], errors="coerce").fillna(1) if fat_col else 1
    )

    agg = (
        ar.groupby("fips")
        .agg(
            fatal_crashes=("fips", "count"),
            total_fatalities=("fatals", "sum"),
            alcohol_fatal_crashes=("alcohol_flag", "sum"),
        )
        .reset_index()
    )
    agg["year"] = year
    return agg


# ======================================================================
#  STEP 1: Extract FARS records for Arkansas, 2008-2023
# ======================================================================

sep("STEP 1: Extracting FARS accident records for Arkansas")

if not FARS_ZIP_DIR.exists():
    log(f"ERROR: FARS_ZIP_DIR '{FARS_ZIP_DIR}' does not exist.")
    log("Create the folder and place all FARS*NationalCSV.zip files inside it.")
    sys.exit(1)

fars_frames = []
for yr in FARS_YEARS:
    result = extract_arkansas(yr)
    if result is not None:
        fars_frames.append(result)

if not fars_frames:
    log("FATAL: No FARS zips found. See instructions at the top of this script.")
    sys.exit(1)

fars_all = pd.concat(fars_frames, ignore_index=True)
fars_all["total_fatalities"]      = fars_all["total_fatalities"].astype(int)
fars_all["alcohol_fatal_crashes"] = fars_all["alcohol_fatal_crashes"].astype(int)

log()
log(f"  Total county-year records extracted: {len(fars_all):,}")
log(f"  Year range: {fars_all['year'].min()}-{fars_all['year'].max()}")
log(f"  Unique county FIPS: {fars_all['fips'].nunique()}")


# ======================================================================
#  STEP 2: Load wet/dry panel from xlsx
# ======================================================================

sep("STEP 2: Loading wet/dry panel from xlsx")

if not XLSX_PATH.exists():
    log(f"ERROR: xlsx not found at '{XLSX_PATH}'. Update XLSX_PATH at the top of this script.")
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

# WetDry_Panel_2010_2024
ws_wd   = wb["WetDry_Panel_2010_2024"]
wd_rows = list(ws_wd.iter_rows(values_only=True))
wd_headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(wd_rows[0])]
wd = pd.DataFrame(wd_rows[1:], columns=wd_headers)
wd["year"] = pd.to_numeric(wd["year"], errors="coerce")
wd["fips"] = wd["fips"].astype(str).str.strip().str.zfill(5)
wd = wd[wd["year"].notna()].copy()
wd["year"] = wd["year"].astype(int)

log(f"  Wet/dry panel rows: {len(wd):,}")
log(f"  Panel years: {wd['year'].min()}-{wd['year'].max()}")
log(f"  Unique counties: {wd['county'].nunique()}")

# Transition_Events
ws_tr    = wb["Transition_Events"]
tr_rows  = list(ws_tr.iter_rows(values_only=True))
tr_headers = [str(h) for h in tr_rows[0]]
transitions = pd.DataFrame(tr_rows[1:], columns=tr_headers)
transitions["fips"]       = transitions["fips"].astype(str).str.strip().str.zfill(5)
transitions["event_year"] = pd.to_numeric(transitions["event_year"], errors="coerce").astype(int)

log(f"  Transition events: {len(transitions)}")

# County_Adjacency_Edges (directed edge list -- used for neighbor identification)
ws_adj   = wb["County_Adjacency_Edges"]
adj_rows = list(ws_adj.iter_rows(values_only=True))
adj_headers = [str(h) for h in adj_rows[0]]
adj = pd.DataFrame(adj_rows[1:], columns=adj_headers)
adj["fips"]          = adj["fips"].astype(str).str.strip().str.zfill(5)
adj["neighbor_fips"] = adj["neighbor_fips"].astype(str).str.strip().str.zfill(5)

log(f"  Adjacency edges loaded: {len(adj):,}")

wb.close()


# ======================================================================
#  STEP 3: Extend wet/dry panel back to 2008-2009
#  (xlsx starts at 2010; backfill using 2010 values,
#   correcting 2010-cohort counties to dry in 2008-2009)
# ======================================================================

sep("STEP 3: Extending wet/dry panel back to 2008-2009")

base_2010       = wd[wd["year"] == 2010].copy()
trans_2010_fips = set(transitions[transitions["event_year"] == 2010]["fips"])

ext_frames = []
for yr in [2008, 2009]:
    tmp = base_2010.copy()
    tmp["year"] = yr
    dry_mask = tmp["fips"].isin(trans_2010_fips)
    for col in ["status_eoy", "status_full_year"]:
        if col in tmp.columns:
            tmp.loc[dry_mask, col] = "dry"
    for col in ["countywide_wet_eoy", "countywide_wet_full_year",
                "wet_any_eoy", "wet_any_full_year"]:
        if col in tmp.columns:
            tmp.loc[dry_mask, col] = 0
    ext_frames.append(tmp)

wd_full = pd.concat(ext_frames + [wd], ignore_index=True)
wd_full = wd_full.sort_values(["fips", "year"]).reset_index(drop=True)

log(f"  Panel rows after extension: {len(wd_full):,}  ({wd_full['year'].min()}-{wd_full['year'].max()})")


# ======================================================================
#  STEP 4: Construct C&S-A cohort identifiers
# ======================================================================

sep("STEP 4: Constructing C&S-A cohort identifiers")

ever_treated_fips = set(transitions["fips"])

cohort_map = (
    transitions
    .sort_values("event_year")
    .drop_duplicates("fips")
    .set_index("fips")["event_year"]
    .to_dict()
)

treated_in_window = {
    fips: yr for fips, yr in cohort_map.items()
    if yr <= ESTIMATION_END
}

fips_to_county = (
    wd_full[["fips", "county"]]
    .drop_duplicates()
    .set_index("fips")["county"]
    .to_dict()
)

log(f"  Ever-treated counties (all years): {len(ever_treated_fips)}")
log(f"  Treated cohorts in window ({ESTIMATION_START}-{ESTIMATION_END}): {len(treated_in_window)}")
log()

for fips, yr in sorted(treated_in_window.items(), key=lambda x: x[1]):
    cname = fips_to_county.get(fips, fips)
    scope = transitions[transitions["fips"] == fips]["policy_scope"].iloc[0]
    pre   = yr - ESTIMATION_START
    post  = ESTIMATION_END - yr
    log(f"    {cname:<16} FIPS={fips}  cohort={yr}  pre={pre}yr  post={post}yr")
    log(f"                       ({scope})")

# Neighbor counties: border a treated county but are not treated themselves
neighbor_fips = (
    set(adj[adj["fips"].isin(ever_treated_fips)]["neighbor_fips"])
    - ever_treated_fips
)

clean_controls = (
    set(wd_full["fips"].unique()) - ever_treated_fips - neighbor_fips
)

log()
log(f"  Neighbor (spillover) counties: {len(neighbor_fips)}")
log(f"  Clean never-treated controls:  {len(clean_controls)}")


# ======================================================================
#  STEP 5: Add treatment/cohort columns to the panel
# ======================================================================

sep("STEP 5: Adding treatment and cohort columns")

wd_full["first_treated_year"] = wd_full["fips"].map(cohort_map)
wd_full["cohort"]             = wd_full["first_treated_year"]

wd_full["years_since_treatment"] = wd_full.apply(
    lambda r: r["year"] - r["first_treated_year"]
    if pd.notna(r["first_treated_year"]) else None,
    axis=1,
)

wd_full["treated_unit"]  = wd_full["fips"].isin(ever_treated_fips).astype(int)
wd_full["neighbor_unit"] = wd_full["fips"].isin(neighbor_fips).astype(int)
wd_full["clean_control"] = wd_full["fips"].isin(clean_controls).astype(int)

# Standardise to original panel column names
wd_full = wd_full.rename(columns={
    "countywide_wet_eoy":                 "countywide_wet",
    "borders_countywide_wet_eoy":         "any_wet_neighbor",
    "n_countywide_wet_neighbors_eoy":     "n_wet_neighbors",
    "share_countywide_wet_neighbors_eoy": "share_wet_neighbors",
    "borders_any_wet_area_eoy":           "any_wet_area_neighbor",
})


# ======================================================================
#  STEP 6: Merge FARS outcomes onto the wet/dry panel
# ======================================================================

sep("STEP 6: Merging FARS outcomes onto wet/dry panel")

wd_window = wd_full[
    (wd_full["year"] >= ESTIMATION_START) &
    (wd_full["year"] <= ESTIMATION_END)
].copy()

merged = wd_window.merge(
    fars_all[["fips", "year", "fatal_crashes", "total_fatalities", "alcohol_fatal_crashes"]],
    on=["fips", "year"],
    how="left",
)

for col in ["fatal_crashes", "total_fatalities", "alcohol_fatal_crashes"]:
    merged[col] = merged[col].fillna(0).astype(int)

merged["alcohol_share"] = merged.apply(
    lambda r: r["alcohol_fatal_crashes"] / r["fatal_crashes"]
    if r["fatal_crashes"] > 0 else 0.0,
    axis=1,
)

zero_crash_rows = (merged["fatal_crashes"] == 0).sum()
matched_rows    = (merged["fatal_crashes"]  > 0).sum()
fips_unmatched  = set(fars_all["fips"]) - set(wd_full["fips"])

log(f"  Merged rows:                       {len(merged):,}")
log(f"  County-years with FARS data:       {matched_rows:,}")
log(f"  County-years with 0 fatal crashes: {zero_crash_rows:,}")
log(f"  FARS FIPS not in wet/dry panel:    {len(fips_unmatched)}"
    + (f"  {fips_unmatched}" if fips_unmatched else "  -- clean merge"))


# ======================================================================
#  STEP 7: Select final columns and sort
# ======================================================================

FINAL_COLS = [
    "fips", "county", "year",
    "fatal_crashes", "total_fatalities", "alcohol_fatal_crashes", "alcohol_share",
    "countywide_wet", "first_treated_year", "cohort", "years_since_treatment",
    "any_wet_neighbor", "n_wet_neighbors", "share_wet_neighbors",
    "any_wet_area_neighbor", "n_neighbors",
    "partial_county", "treated_unit", "neighbor_unit", "clean_control",
]

final_cols = [c for c in FINAL_COLS if c in merged.columns]
panel_out  = merged[final_cols].sort_values(["fips", "year"]).reset_index(drop=True)


# ======================================================================
#  STEP 8: Build updated transition summary
# ======================================================================

trans_out = transitions.copy()
trans_out["in_fars_window"]        = (trans_out["event_year"] <= ESTIMATION_END).astype(int)
trans_out["pre_periods_available"] = trans_out["event_year"] - ESTIMATION_START


# ======================================================================
#  DESCRIPTIVE SUMMARY
# ======================================================================

sep(f"DESCRIPTIVE SUMMARY ({ESTIMATION_START}-{ESTIMATION_END})")

est   = panel_out
n_yrs = ESTIMATION_END - ESTIMATION_START + 1

log(f"  Obs: {len(est):,}  ({est['fips'].nunique()} counties x up to {n_yrs} years)")
log(f"  Treated units:    {panel_out[panel_out['treated_unit']==1]['fips'].nunique()} counties")
log(f"  Neighbor units:   {panel_out[panel_out['neighbor_unit']==1]['fips'].nunique()} counties")
log(f"  Clean controls:   {panel_out[panel_out['clean_control']==1]['fips'].nunique()} counties")
log()
log(f"  Fatal crashes   (mean per county-year):  {est['fatal_crashes'].mean():.2f}")
log(f"  Alcohol crashes (mean per county-year):  {est['alcohol_fatal_crashes'].mean():.2f}")
log(f"  Avg alcohol share:                       {est['alcohol_share'].mean():.3f}")
log()
log(f"  Treated cohorts in window (cohort <= {ESTIMATION_END}):")
for fips, yr in sorted(treated_in_window.items(), key=lambda x: x[1]):
    cname = fips_to_county.get(fips, fips)
    pre   = yr - ESTIMATION_START
    post  = ESTIMATION_END - yr
    log(f"    {cname:<16} cohort={yr}  pre={pre}yr  post={post}yr")

log()
treat_pre  = panel_out[(panel_out["treated_unit"]==1) & (panel_out["years_since_treatment"] < 0)]
treat_post = panel_out[(panel_out["treated_unit"]==1) & (panel_out["years_since_treatment"] >= 0)]
log(f"  Pre-treatment alcohol share (treated):   {treat_pre['alcohol_share'].mean():.3f}")
log(f"  Post-treatment alcohol share (treated):  {treat_post['alcohol_share'].mean():.3f}")


# ======================================================================
#  STEP 9: Export
# ======================================================================

sep("STEP 9: Exporting outputs")

OUT_DIR.mkdir(parents=True, exist_ok=True)

panel_path = OUT_DIR / "arkansas_did_panel.csv"
trans_path = OUT_DIR / "transition_summary.csv"
log_path   = OUT_DIR / "merge_log.txt"

panel_out.to_csv(panel_path, index=False)
trans_out.to_csv(trans_path, index=False)
log_path.write_text("\n".join(LOG_LINES))

log()
log(f"  checkmark  Main panel saved:         {panel_path}")
log(f"    Rows: {len(panel_out):,}  |  Columns: {len(panel_out.columns)}")
log(f"    Year range: {panel_out['year'].min()}-{panel_out['year'].max()}")
log()
log(f"  checkmark  Transition summary saved: {trans_path}")
log(f"  checkmark  Log saved:                {log_path}")
